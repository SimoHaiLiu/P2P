import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
import json
import xlrd, xlwt
import pandas as pd
from xlutils.copy import copy
import os
from openpyxl import load_workbook
import openpyxl
import datetime
import shutil
import pdfplumber
import pymongo
import re
from selenium.webdriver.chrome.options import Options

login_url = 'https://app.qupital.com/login'
root_url = 'https://app.qupital.com/funder/involved/past'
base_url = 'https://app.qupital.com'

cookies_str = '__guid=238974800.2252233793311927000.1540203821649.4636; galaxy-sticky=!MKpDKsP5zQveCErSr-c97x5; monitor_count=4; x_mtok=hszhgHT5ToA7GvPLW'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.26 Safari/537.36 Core/1.63.6756.400 QQBrowser/10.3.2473.400',
    'Cookie': cookies_str}
post_data = {'loginEmail': 'christina.lee@shkf.com', 'loginPassword': 'chasiu456'}

seller_report_path = 'seller_report'
cookies_file = 'qupital_cookies.json'
local_storage_file = 'qupital_local_storage.json'
# 是否下载卖家报告的开关
download_pdf = True

mongo_client = pymongo.MongoClient('localhost', 27017)
db = mongo_client.qupital


def login(driver):
    driver.get(login_url)
    wait = WebDriverWait(driver, 20)

    login_email = wait.until(EC.presence_of_element_located((By.ID, 'loginEmail')))
    login_email.send_keys(post_data['loginEmail'])
    login_password = wait.until(EC.presence_of_element_located((By.ID, 'loginPassword')))
    login_password.send_keys(post_data['loginPassword'])
    login_button = wait.until(EC.presence_of_element_located((By.ID, 'loginButton')))
    login_button.click()
    time.sleep(5)

    return driver


def save_token(driver, token_type='cookies'):
    '''
    保存cookies或者local_storage
    :param token: cookies或者local_storage字典数据
    :param token_type: cookies或者local_storage
    :return:
    '''
    if token_type == 'cookies':
        token_file = 'qupital_cookies.json'
        token = driver.get_cookies()
    elif token_type == 'local_storage':
        token_file = 'qupital_local_storage.json'
        token = {'Meteor.loginToken': driver.execute_script('return localStorage.getItem("Meteor.loginToken")'),
                 'Meteor.loginTokenExpires': driver.execute_script(
                     'return localStorage.getItem("Meteor.loginTokenExpires")'),
                 'Meteor.userId': driver.execute_script('return localStorage.getItem("Meteor.userId")')}

    with open(token_file, 'w', encoding='utf-8') as file:
        file.write(json.dumps(token, indent=4, ensure_ascii=False))


def read_token(token_type='cookies'):
    '''
    读取cookies或者local_storage文件
    :return:
    '''
    if token_type == 'cookies':
        token_file = 'qupital_cookies.json'
    elif token_type == 'local_storage':
        token_file = 'qupital_local_storage.json'

    with open(token_file, 'r', encoding='utf-8') as file:
        token = json.loads(file.read())

    if token_type == 'cookies':
        # 拼接出请求头的cookies
        # return '; '.join([item['name'] + '=' + item['value'] for item in cookies])
        return {item['name']: item['value'] for item in token}
    elif token_type == 'local_storage':
        return token


def pdf_data_extract(filename, currency):
    '''
    从PDF中提取卖家报告数据
    :param filename:
    :return:
    '''
    # 卖家报告
    seller_report = {'latest_share_capital': None, 'second_latest_share_capital': None, 'client': None,
                     'establishment_of_relationship': None, 'revenue': None, 'revenue_currency': None,
                     'trade_receivables': None, 'cash_and_bank_balance': None, 'bank_borrowing': None,
                     'total_equity': None, 'gearing_ratio': None}
    # Client chosen for financing和Years of relationship存在多条数据，暂定为只读取第一条
    # client_first_read = True
    # relationship_first_read = True

    # 一个字段存在多张表多个年份的数据
    share_capital = []
    revenue = []
    trade_receivables = []
    cash_and_bank_balance = []
    bank_borrowing = []
    total_equity = []
    gearing_ratio = []
    # 存在多个
    client = []
    establishment_of_relationship = []

    pdf = pdfplumber.open(filename)
    for page in pdf.pages:
        # 获取当前页面的全部文本信息，包括表格中的文字
        # print(page.extract_text())
        # continue

        for table in page.extract_tables():
            for row in table:
                row = cleaning_row(row)
                # 因为字段的大小写和空格数都不确定，所以转成小写并去除所有空格方便查找
                join_row = ''.join(row).lower().replace(' ', '')
                if 'share capital' in row:
                    print(row)
                    # 最新股本
                    # seller_report['latest_share_capital'] = nil_check(row[1])
                    # 前一年股本
                    # seller_report['second_latest_share_capital'] = nil_check(row[2])
                    merge_list(row, share_capital)
                elif 'clientchosenfor' in join_row:
                    # 有的字段名为Client Chosen for Financing，有的字段名为Client Chosen for Invoice Exchange
                    print(row)
                    # 卖家客户
                    merge_list(row, client)
                elif 'establishmentofrelationship-when?' in join_row or 'yearsofrelationship' in join_row:
                    # 有的叫Establishment  of \nrelationship - When?，有的叫Years of relationship
                    print(row)
                    # 合作关系
                    merge_list(row, establishment_of_relationship)
                elif 'revenue' in row:
                    print(row)
                    # 卖家盈利
                    # seller_report['revenue'] = nil_check(row[1])
                    merge_list(row, revenue)
                elif 'currency' in row:
                    print(row)
                    # 暂时无法识别那个Currency没有横杠
                    # 卖家盈利货币 Todo: 是否为相同的Currency需要确认
                    # seller_report['revenue_currency'] = currency
                elif 'trade receivables' in row:
                    print(row)
                    # 应收贸易账项
                    # seller_report['trade_receivables'] = nil_check(row[1])
                    merge_list(row, trade_receivables)
                elif 'cash and bank balance' in row:
                    print(row)
                    # 现金及银行结余
                    # seller_report['cash_and_bank_balance'] = nil_check(row[1])
                    merge_list(row, cash_and_bank_balance)
                elif 'bank borrowing' in row or 'bank borrowings' in row:
                    # 有的PDF没有该字段，例如：AU12889646.pdf
                    print(row)
                    # 银行借款
                    # seller_report['bank_borrowing'] = nil_check(row[1])
                    merge_list(row, bank_borrowing)
                elif 'total equity' in row:
                    print(row)
                    # 股本
                    # seller_report['total_equity'] = nil_check(row[1])
                    merge_list(row, total_equity)
                elif 'gearing ratio' in row:
                    print(row)
                    # 杠杆比率
                    # seller_report['gearing_ratio'] = nil_check(row[1])
                    merge_list(row, gearing_ratio)

    # print('Share Capital All: {}'.format(share_capital))
    seller_report['latest_share_capital'] = nil_check(share_capital, -1)
    seller_report['second_latest_share_capital'] = nil_check(share_capital, -2)
    seller_report['revenue'] = nil_check(revenue, -1)
    seller_report['trade_receivables'] = nil_check(trade_receivables, -1)
    seller_report['cash_and_bank_balance'] = nil_check(cash_and_bank_balance, -1)
    seller_report['bank_borrowing'] = nil_check(bank_borrowing, -1)
    seller_report['total_equity'] = nil_check(total_equity, -1)
    seller_report['gearing_ratio'] = nil_check(gearing_ratio, -1)
    if len(client) == 1:
        seller_report['client'] = client[0]
    elif len(client) > 1:
        seller_report['client'] = client
    if len(establishment_of_relationship) == 1:
        seller_report['establishment_of_relationship'] = establishment_of_relationship[0]
    elif len(establishment_of_relationship) > 1:
        seller_report['establishment_of_relationship'] = establishment_of_relationship
    # seller_report['currency'] = currency

    print(seller_report)
    return seller_report


def nil_check(value, index=0):
    '''
    PDF文件中Balance Sheet表中存在空值，例如：AU12889699.pdf
    :param value:
    :return:
    '''
    # print('**** {} ****'.format(value))
    if isinstance(value, list):
        if len(value) == 0:
            return None
        # 如果索引超范围则直接返回空值
        elif index < 0 and len(value) < abs(index):
            return None
        elif index > 0 and len(value) < index + 1:
            return None
        else:
            return None if value[index] == 'Nil' or value[index] == 'N/A' else str_convert_num(value[index])
    else:
        return None if value == 'Nil' or value == 'N/A' else str_convert_num(value)


def merge_list(row, field_list):
    # 当只有表头数据时不做处理，例如：['Years of relationship']
    if len(row) > 1:
        # 因为时间是2016 -> 2015这样排序的，所以需要去除表头后反转列表
        if row[0] == 'client chosen for financing' or row[0] == 'years of relationship' or row[0].replace(' ',
                                                                                                          '') == 'establishmentofrelationship-when?':
            row = row[1:]
        else:
            row = row[1:][::-1]

        field_list.extend(row)


def cleaning_row(row):
    temp_row = []
    # 去除可能包含的None值
    row = list(filter(None, row))

    # 去除换行符和多余的空格
    for index, item in enumerate(row):
        if '\n' in item:
            item = item.replace('\n', '')

        if index == 0:
            # row[index] = item.lower()
            # 因为PDF字段有的大小写不确定，所以将所以的表头转成小写便于判断
            temp_row.append(item.lower())
            continue

        # 将多个替换成一个空格
        item = re.sub(' +', ' ', item)

        # 这些字段数据有可能类似(982,190)，所以需要去掉小括号
        if len(row) > 1 and (temp_row[0] == 'total equity' or temp_row[0] == 'cash and bank balance' or temp_row[
            0] == 'bank borrowings' or temp_row[0] == 'bank borrowing' or temp_row[0] == 'gearing ratio' or temp_row[
                                 0] == 'revenue' or temp_row[0] == 'trade receivables'):
            item = re.sub('\(|\)', '', item)
        else:
            # 将类似1)、2)的字符串去除，只匹配一次
            item = re.sub('\d+\)', '', item, 1)
        # item = item.lower()
        temp_row.append(item.strip())

    # 去除空字符串并返回
    return list(filter(None, temp_row))


# 获取页面所有的 Auction 链接
def new_download_file():
    '''
    获取最新的下载文件
    :return:
    '''
    return max([os.path.join(seller_report_path, f) for f in os.listdir(seller_report_path)], key=os.path.getctime)


def rename_pdf(auction_no):
    '''
    以Auction No为文件名重命名PDF文件
    :param auction_no:
    :return:
    '''
    filename = new_download_file()
    new_filename = os.path.join(seller_report_path, auction_no + '.pdf')

    # 当文件还在下载的时候不能重命名
    while os.path.splitext(filename)[1] != '.pdf':
        time.sleep(1)
        filename = new_download_file()

    # shutil.move(filename, os.path.join(seller_report_path, auction_no + '.pdf'))

    for retry in range(3):
        if os.path.exists(new_filename):
            try:
                # 当文件处于打开状态删除的时候会报PermissionError
                os.remove(new_filename)
            except Exception as e:
                print('删除文件报错：{}'.format(e))
                print('文件可能被其他应用程序打开，尝试关闭文件')
                time.sleep(3)
        else:
            break

    try:
        os.rename(filename, new_filename)
        print('{} rename to {}'.format(filename, new_filename))
        return new_filename
    except Exception as e:
        print('重命名文件报错：{}'.format(e))
        return filename

    # return filename


def str_convert_num(value):
    '''
    将字符串格式化成数字
    :param value:
    :return:
    '''
    if ',' in value:
        value = value.replace(',', '')

    if '.' in value:
        if value.replace('.', '').isdigit():
            return float(value)
        else:
            return None
    else:
        if value.isdigit():
            return int(value)
        else:
            return None


def save_to_mongo(collection, auction_dict):
    auction_info = db[collection].find_one({'auction_no': auction_dict['auction_no']})
    print('auction_info: '.format(auction_info))

    if auction_info:
        # if collection == 'current_auctions':
        #     print('update')
        #     db[collection].update_one({'auction_no': basic_information['auction_no']}, {
        #         '$set': dict(basic_information, **seller_statistic, **seller_obligor_statistic, **seller_report)})
        # elif collection == 'past_auctions':
        #     db[collection].update_one({'auction_no': basic_information['auction_no']}, {
        #         '$set': dict(auction_record_table, **basic_information, **seller_statistic, **seller_obligor_statistic,
        #                      **seller_report)})
        db[collection].update_one({'auction_no': auction_dict['auction_no']}, {
            '$set': auction_dict})
    else:
        # if collection == 'current_auctions':
        #     print('insert')
        #     db[collection].insert_one(
        #         dict(basic_information, **seller_statistic, **seller_obligor_statistic,
        #              **seller_report))
        # elif collection == 'past_auctions':
        #     db[collection].insert_one(
        #         dict(auction_record_table, **basic_information, **seller_statistic, **seller_obligor_statistic,
        #              **seller_report))
        db[collection].insert_one(auction_dict)


def next_page(driver):
    # 判断有没有下一页
    try:
        next_page = driver.find_element_by_xpath('//li[@title="next page"]/a')
        next_page.click()
        return driver
    except Exception as e:
        print('next_page: {}'.format(e))
        print('没有下一页了')
        # driver.close()
        driver.quit()


def auction_detail(driver, auction_type='current_auctions'):
    if auction_type == 'past_auctions':
        auctions = driver.find_elements_by_xpath('//a[contains(@href, "/funder/auction")]')
        auctions_table_tr = driver.find_elements_by_xpath('//table[@class="table table-striped table-bordered"]//tr')

        for auction, auction_table_tr in zip(auctions, auctions_table_tr):
            auction_record_table = {}

            auctions_table_td = auction_table_tr.find_elements_by_xpath('./td[not(contains(@style, "display: none"))]')
            auction_record_table['net_total_to_be_received'] = nil_check(auctions_table_td[3].text)
            auction_record_table['net_return'] = nil_check(auctions_table_td[4].text)
            auction_record_table['advanced_date'] = datetime.datetime.strptime(auctions_table_td[5].text, '%Y-%m-%d')
            auction_record_table['remitted_date'] = None if auctions_table_td[
                                                                6].text == 'NA' else datetime.datetime.strptime(
                auctions_table_td[6].text, '%Y/%m/%d')
            auction_record_table['late_day'] = nil_check(auctions_table_td[7].text)
            auction_record_table['status'] = auctions_table_td[8].text
            print(auction_record_table)

            if download_pdf:
                basic_information, seller_statistic, seller_obligor_statistic, seller_report = auction_detail_extract(
                    driver, auction)
                save_to_mongo(auction_type, dict(basic_information, **seller_statistic, **seller_obligor_statistic, **seller_report, **auction_record_table))
            else:
                basic_information, seller_statistic, seller_obligor_statistic = auction_detail_extract(
                    driver, auction)
                save_to_mongo(auction_type, dict(basic_information, **seller_statistic, **seller_obligor_statistic, **auction_record_table))
    elif auction_type == 'current_auctions':
        auctions = driver.find_elements_by_xpath('//div[@class="cardblock"]/a')

        for auction in auctions:
            if download_pdf:
                basic_information, seller_statistic, seller_obligor_statistic, seller_report = auction_detail_extract(driver, auction)
                save_to_mongo(auction_type, dict(basic_information, **seller_statistic, **seller_obligor_statistic, **seller_report))
            else:
                basic_information, seller_statistic, seller_obligor_statistic = auction_detail_extract(driver, auction)
                save_to_mongo(auction_type, dict(basic_information, **seller_statistic, **seller_obligor_statistic))
            print('#' + '-' * 30 + '#')


def auction_detail_extract(driver, auction):
    # 基本信息
    basic_information = {}
    # 买家数据
    seller_statistic = {}
    # 买卖方数据
    seller_obligor_statistic = {}
    # 判断 key 是否已经存在
    keys_set = set()

    href = auction.get_attribute('href')
    print(href)

    # 打开 Auction 详情窗口
    js = 'window.open("' + href + '");'
    driver.execute_script(js)
    handles = driver.window_handles
    driver.switch_to.window(handles[-1])
    # 等待js渲染完成，否则有些数据没有加载出来
    time.sleep(5)

    discuss_div = driver.find_elements_by_xpath('//div[@class="discuss"]/p')
    # 发票编号
    auction_no = discuss_div[0].text.strip().replace('Auction no.: ', '')
    # print('auction_no: {}'.format(auction_no))
    basic_information['auction_no'] = auction_no

    # 如果本地已经下载了卖家报告，则无需重复下载
    if os.path.exists(os.path.join(seller_report_path, auction_no+'.pdf')):
        print('{} Seller Report is already download!'.format(auction_no))
    else:
        # 下载Seller Report
        print('下载Seller Report')
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//p/a')))
        print(driver.find_elements_by_xpath('//p/a')[1].text)
        driver.find_elements_by_xpath('//p/a')[1].click()
        time.sleep(5)

    # 信贷保险
    if discuss_div[1].find_element_by_xpath('./span').get_attribute('class') == 'glyphicon glyphicon-remove':
        print('credit_insurance: {}'.format(0))
        basic_information['credit_insurance'] = 0
    elif discuss_div[1].find_element_by_xpath('./span').get_attribute('class') == 'glyphicon glyphicon-ok':
        print('credit_insurance: {}'.format(1))
        basic_information['credit_insurance'] = 1
    # 债务人知情
    if discuss_div[2].find_element_by_xpath('./span').get_attribute('class') == 'glyphicon glyphicon-remove':
        print('obligor_notification: {}'.format(0))
        basic_information['obligor_notification'] = 0
    elif discuss_div[2].find_element_by_xpath('./span').get_attribute('class') == 'glyphicon glyphicon-ok':
        print('obligor_notification: {}'.format(1))
        basic_information['obligor_notification'] = 1

    h3_text = driver.find_elements_by_xpath('//h3')[0].text.strip().split('\n')  # .replace('Obligor name: ', '')
    # 债务人名称
    obligor_name = h3_text[0].replace('Obligor name: ', '')
    print('obligor_name: {}'.format(obligor_name))
    basic_information['obligor_name'] = obligor_name
    # 余下时间
    remaining_time = h3_text[1].replace('Remaining time: ', '')
    print('remaining_time: {}'.format(remaining_time))
    basic_information['remaining_time'] = remaining_time

    h4_text = driver.find_elements_by_xpath('//h4')[0].text.strip().split(' ')
    # 货币
    # print(h4_text)
    currency = h4_text[0]
    print('currency: {}'.format(currency))
    basic_information['currency'] = currency
    # 预付金额
    invoice_amount = str_convert_num(h4_text[1])
    print('invoice_amount: {}'.format(invoice_amount))
    basic_information['invoice_amount'] = invoice_amount
    # 结账天数
    print('no_of_days: {}'.format(str_convert_num(h4_text[3])))
    basic_information['no_of_days'] = str_convert_num(h4_text[3])

    keys = driver.find_elements_by_xpath('//strong')[0:32]
    for k in keys:
        value = k.find_element_by_xpath('./parent::*').text

        key = k.text
        value = value.replace(key, '').strip()
        key = key.replace(':', '').strip()

        if key == 'Seller No.':  #
            # 卖家编号
            print('seller_no: {}'.format(value))
            basic_information['seller_no'] = value
        elif key == 'Suggested Basis Point':
            # 回报率
            print('suggested_basis_point: {}'.format(str_convert_num(value)))
            basic_information['suggested_basis_point'] = str_convert_num(value)
        elif key == 'Seller Qupital rating':
            # 卖家信贷评级
            print('seller_qupital_rating: {}'.format(value))
            basic_information['seller_qupital_rating'] = value
        elif key == 'Obligor Qupital rating':
            # 债务人信贷评级
            print('obligor_qupital_rating: {}'.format(str_convert_num(value)))
            basic_information['obligor_qupital_rating'] = str_convert_num(value)
        elif key == 'Expected payment date':
            # 预计付款日期
            print('expected_payment_date: {}'.format(datetime.datetime.strptime(value, '%Y-%m-%d')))
            basic_information['expected_payment_date'] = datetime.datetime.strptime(value, '%Y-%m-%d')
        elif key == 'Buy now annualized yield':
            # 直接购买发票的回报
            print('buy_now_annualized_yield: {}'.format(str_convert_num(value.replace(' basis points', ''))))
            basic_information['buy_now_annualized_yield'] = str_convert_num(value.replace(' basis points', ''))
        # ----------------- Seller Statistic -----------------#
        elif key == 'Successful auction' and 'Successful auction' not in keys_set:
            # 成功发出的发票
            print('ss_successful_auction: {}'.format(str_convert_num(value)))
            seller_statistic['ss_successful_auction'] = str_convert_num(value)
        elif key == 'Remitted auction' and 'Remitted auction' not in keys_set:
            # 已汇款的发票
            print('ss_remitted_auction: {}'.format(str_convert_num(value)))
            seller_statistic['ss_remitted_auction'] = str_convert_num(value)
        elif key == 'Outstanding auction' and 'Outstanding auction' not in keys_set:
            # 未支付的发票
            print('ss_outstanding_auction: {}'.format(str_convert_num(value)))
            seller_statistic['ss_outstanding_auction'] = str_convert_num(value)
        elif key == 'Auction total invoice amount in USD' and 'Auction total invoice amount in USD' not in keys_set:
            # 需要偿还的总数（美元）
            print('ss_auction_total_invoice_amount_in_usd: {}'.format(str_convert_num(value)))
            seller_statistic['ss_auction_total_invoice_amount_in_usd'] = str_convert_num(value)
        elif key == 'Auction total advance amount in USD' and 'Auction total advance amount in USD' not in keys_set:
            # 已收到的预付总数（美元）
            print('ss_auction_total_advance_amount_in_usd: {}'.format(str_convert_num(value)))
            seller_statistic['ss_auction_total_advance_amount_in_usd'] = str_convert_num(value)
        elif key == 'Auction total invoice amount in HKD' and 'Auction total invoice amount in HKD' not in keys_set:
            # 需要偿还的总数（港元）
            print('ss_auction_total_invoice_amount_in_hkd: {}'.format(str_convert_num(value)))
            seller_statistic['ss_auction_total_invoice_amount_in_hkd'] = str_convert_num(value)
        elif key == 'Auction total advance amount in HKD' and 'Auction total advance amount in HKD' not in keys_set:
            # 已收到的预付总数（港元）
            print('ss_auction_total_advance_amount_in_hkd: {}'.format(str_convert_num(value)))
            seller_statistic['ss_auction_total_advance_amount_in_hkd'] = str_convert_num(value)
        elif key == 'Average late day' and 'Average late day' not in keys_set:
            # 平均迟还天数
            print('ss_average_late_day: {}'.format(str_convert_num(value)))
            basic_information['ss_average_late_day'] = str_convert_num(value)
            # seller_statistic['average_late_day'] = str_convert_num(value)
        # ------------------ 买卖方数据 ------------------#
        elif key == 'Successful auction':
            # 成功发出的发票
            print('sos successful_auction: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_successful_auction'] = str_convert_num(value)
        elif key == 'Remitted auction':
            # 已汇款的发票
            print('sos remitted_auction: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_remitted_auction'] = str_convert_num(value)
        elif key == 'Outstanding auction':
            # 未支付的发票
            print('sos outstanding_auction: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_outstanding_auction'] = str_convert_num(value)
        elif key == 'Auction total invoice amount in USD':
            # 需要偿还的总数（美元）
            print('sos auction_total_invoice_amount_in_usd: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_auction_total_invoice_amount_in_usd'] = str_convert_num(value)
        elif key == 'Auction total advance amount in USD':
            # 已收到的预付总数（美元）
            print('sos auction_total_advance_amount_in_usd: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_auction_total_advance_amount_in_usd'] = str_convert_num(value)
        elif key == 'Auction total invoice amount in HKD':
            # 需要偿还的总数（港元）
            print('sos auction_total_invoice_amount_in_hkd: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_auction_total_invoice_amount_in_hkd'] = str_convert_num(value)
        elif key == 'Auction total advance amount in HKD':
            # 已收到的预付总数（港元）
            print('sos auction_total_advance_amount_in_hkd: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_auction_total_advance_amount_in_hkd'] = str_convert_num(value)
        elif key == 'Average late day':
            # 平均迟还天数
            print('sos average_late_day: {}'.format(str_convert_num(value)))
            seller_obligor_statistic['sos_average_late_day'] = str_convert_num(value)

        keys_set.add(key)


    # 关闭当前窗口
    driver.close()
    # 切换到第一个页面
    driver.switch_to.window(handles[0])

    if download_pdf:
        pdf_file = rename_pdf(auction_no)
        seller_report = pdf_data_extract(pdf_file, currency)
        return basic_information, seller_statistic, seller_obligor_statistic, seller_report
    else:
        return basic_information, seller_statistic, seller_obligor_statistic


def main(url):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Runs Chrome in headless mode.
    options.add_argument('--no-sandbox')  # Bypass OS security model
    options.add_argument('--disable-gpu')  # applicable to windows os only
    # options.add_argument('start-maximized')  #
    # options.add_argument('disable-infobars')
    # options.add_argument("--disable-extensions")
    # options.add_argument('--ignore-gpu-blacklist')
    # options.add_argument('--log-level=3')
    # options.add_argument('--disable-dev-shm-usage')

    # options.add_argument('--ignore-gpu-blacklist')
    # options.add_argument('--no-default-browser-check')
    # options.add_argument('--no-first-run')
    # options.add_argument('--disable-default-apps')
    # options.add_argument('--disable-infobars')
    # options.add_argument('--disable-extensions')
    # options.add_argument('--test-type')

    # options.add_argument('--disable-web-security')
    # options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-popup-blocking')

    # options.add_argument('--disable-dev-shm-usage')

    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")
    options.add_argument("--ignore-certificate-errors")
    prefs = {
        "download.default_directory": os.path.join(os.getcwd(), seller_report_path),
        "download.prompt_for_download": False, # 允许自动下载
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True, # allow download of pdf instead of open in plugin
        'safebrowsing.enabled': False, # allow download of .msi, .exe files, etc.
        'safebrowsing.disable_download_protection': True,
        'plugins.plugins_disabled': ["Chrome PDF Viewer"]
    }
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(options=options)

    # $sessionId
    driver.command_executor._commands["send_command"] = ("POST", '/session/{}/chromium/send_command'.format(driver.session_id))
    params = {'cmd': 'Page.setDownloadBehavior',
              'params': {'behavior': 'allow', 'downloadPath': os.path.join(os.getcwd(), seller_report_path)}}
    command_result = driver.execute("send_command", params)
    print('driver.session_id: {}'.format(driver.session_id))
    for key in command_result:
        print('{}: {}'.format(key, command_result[key]))

    if os.path.exists(local_storage_file):
        # 必须要先加载一下网站，不然直接添加cookies会报错
        driver.get(login_url)
        # for name, value in read_token().items():
        #     driver.add_cookie({'name': name, 'value': value})

        token = read_token(token_type='local_storage')
        print(token['Meteor.loginToken'])
        driver.execute_script(
            'localStorage.setItem("{}", "{}")'.format('Meteor.loginToken', token['Meteor.loginToken']))
        driver.execute_script(
            'localStorage.setItem("{}", "{}")'.format('Meteor.loginTokenExpires', token['Meteor.loginTokenExpires']))
        driver.execute_script('localStorage.setItem("{}", "{}")'.format('Meteor.userId', token['Meteor.userId']))
    else:
        driver = login(driver)
        save_token(driver)
        save_token(driver, token_type='local_storage')

    # url = 'https://app.qupital.com/funder/involved/past'
    driver.get(url)
    time.sleep(5)
    if url == 'https://app.qupital.com/funder/involved/past':
        # -----------------------暂时不需要-----------------------#
        # web_driver_wait(driver, EC.presence_of_element_located((By.XPATH, '//input[@class="form-control"]')), '信息页面')
        # start_date = driver.find_elements_by_xpath('//input[@class="form-control"]')[0]
        # start_date.clear()
        # start_date.send_keys('2018/01/01')

        # web_driver_wait(driver, EC.presence_of_element_located((By.CLASS_NAME, 'btn-primary')))
        # search_button = driver.find_element_by_class_name('btn-primary')
        # search_button.click()
        # 等待Search的请求完成
        # wait.until_not(EC.presence_of_element_located((By.CLASS_NAME, 'ReactModal__Body--open')))
        # -----------------------暂时不需要-----------------------#

        # select_button = Select(driver.find_element_by_class_name('select-filter'))
        # select_button.select_by_index(0)
        #
        while driver != None:
            auction_detail(driver, 'past_auctions')
            driver = next_page(driver)
    elif url == 'https://app.qupital.com/funder/list':
        while True:
            auction_detail(driver)
            print('持续监测')
            time.sleep(1)

############################# 文件相关操作 #############################
# def write_dict_to_file():
#     datas = main()
#     with open('../data/qupital/qupital-v2.txt', 'w', encoding='utf-8') as f:
#         data = json.dumps(datas)
#         f.write(data)
#
#
# def read_dict_from_file():
#     with open('../data/qupital/qupital-v2.txt', 'r', encoding='utf-8') as file:
#         data = file.readline()
#         # print(type(json.loads(data)))
#         return json.loads(data)
#
#
# def format_file():
#     json_data = read_dict_from_file()
#     with open('../data/qupital/format_qupital-v2.txt', 'w', encoding='utf-8') as file:
#         data = json.dumps(json_data, indent=2)
#         file.write(data)
#
#
# def read_excel():
#     data = xlrd.open_workbook('../data/qupital/Qupital-FN12180105-2018-10-22-Auction-Summary (1).xls')
#     table = data.sheets()[0]
#     nrows = table.nrows
#     ncols = table.ncols
#
#     for i in range(nrows):
#         print(table.row_values(i))
#
#
# def merge_csv():
#     data = read_dict_from_file()
#     df = pd.read_csv('../data/qupital/Qupital-FN12180105-2018-10-22-Auction-Summary.csv')
#
#     data_df = pd.DataFrame(list(data.values()), index=list(data.keys()))
#     data_df['Auction no.'] = data_df.index
#
#     merge_data = pd.merge(df, data_df, on='Auction no.')
#     merge_data.to_csv('../data/qupital/full_Qupital-FN12180105-2018-10-22-Auction-Summary.csv', index=False)
#
#
# def data_clean():
#     df = pd.read_csv('../data/qupital/full_Qupital-FN12180105-2018-10-22-Auction-Summary.csv')
#     print(df)
#     print(df['Minimum advance amount'].str.contains('USD'))


# 补充缺少的 SS 数据
# def set_ss_data():
#     excel_file = '../data/qupital/20181024 invoice testing v2.xlsx'
#     # test_excel_file = '../data/qupital/test-excel-file.xlsx'
#     ss_file = '../data/qupital/format_qupital-v2.txt'
#     # 待填充的列
#     ss_columns = ['SS Auction total invoice amount in HKD', 'SS Auction total invoice amount in USD',
#                   'SS Average late day', 'SS Outstanding auction', 'SS Remitted auction']
#     # 待填充的列的列数
#     ss_column_num = {}
#
#     with open(ss_file, 'r', encoding='utf-8') as file:
#         ss_data = json.load(file)
#     # print(ss_data)
#
#     wb = load_workbook(excel_file)
#     sheet = wb.active
#     # print(sheet.cell(2, 2).value)
#     # wb.save(test_excel_file)
#
#     # excel 列数
#     max_column = sheet.max_column
#     # excel 行数
#     max_row = sheet.max_row
#
#     for column in range(1, max_column + 1):
#         # 获取表头
#         column_value = sheet.cell(1, column).value
#         if column_value in ss_columns:
#             ss_column_num[column_value] = column
#
#     # 获取爬取的数据所有的 AU 键
#     ss_data_keys = ss_data.keys()
#     for row in range(2, max_row + 1):
#         au_id = sheet.cell(row, 1).value
#         # 在爬取的数据中查找 AU 键
#         # for ss_data_key in ss_data.keys():
#         if au_id in ss_data_keys:
#             print(au_id)
#             for column_name, column_num in ss_column_num.items():
#                 sheet.cell(row, column_num, ss_data[au_id][column_name])
#                 # print(column_name)
#
#     wb.save(excel_file)


############################# 获取Cookies，直接登录 #############################
# def login_with_cookie():
#     headers = {
#         'Cookie': 'galaxy-sticky=!MKpDKsP5zQveCErSr-4nxvw; x_mtok=f8B7YbDqyvBoRzz2D',
#         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
#     }
#     url = 'https://app.qupital.com/funder/involved/past'
#     session = requests.Session()
#     response = session.get(url, headers=headers)
#     print(response.status_code)
#     print(response.text)


if __name__ == '__main__':
    # file_path = '../data/qupital/Qupital-FN12180105-2018-10-22-Auction-Summary (1).xls'
    # send_request()
    # webdriver_login()

    # 爬取数据并保存到本地文件
    # write_dict_to_file()
    # read_dict_from_file()
    # format_file()

    # 将爬取的数据和下载的数据合并
    # merge_csv()

    '''
    处理数据：
    1. 1054 basis point —> 10.54  （除100）
    2. Usd  19800 -> 19800
    '''
    # data_clean()

    # set_ss_data()

    if not os.path.exists(seller_report_path):
        os.makedirs(seller_report_path)

    # 当前可拍卖发票
    # main('https://app.qupital.com/funder/list')
    # 历史拍卖纪录
    main('https://app.qupital.com/funder/involved/past')

    # PDF内容提取测试
    # pdf_data_extract(os.path.join(seller_report_path, 'AU12890363.pdf'), 233333)
